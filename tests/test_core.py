from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.core import (
    carrier_domain_counts, export_recipient_file, load_recipient_file,
    match_individual_attachments, render_template, split_addresses,
    typo_domain_suspects, unknown_tags, validate_rows,
)


def test_load_xlsx_and_normalize(tmp_path: Path):
    path = tmp_path / "data.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.append(["事業所名", "人数", "メール"])
    sheet.append(["○○商事", 2, "a@example.jp"])
    book.save(path)
    result = load_recipient_file(str(path))
    assert result.headers == ["事業所名", "人数", "メール"]
    assert result.rows[0]["人数"] == "2"


def test_export_recipient_file_writes_headers_and_rows(tmp_path: Path):
    path = tmp_path / "out.xlsx"
    headers = ["氏名", "メール"]
    rows = [
        {"氏名": "山田", "メール": "a@example.jp"},
        {"氏名": "佐藤", "メール": "b@example.jp"},
    ]
    export_recipient_file(str(path), headers, rows)
    book = load_workbook(path)
    sheet = book.active
    values = [list(row) for row in sheet.iter_rows(values_only=True)]
    assert values == [
        ["氏名", "メール"],
        ["山田", "a@example.jp"],
        ["佐藤", "b@example.jp"],
    ]


def test_render_keeps_unknown_tags():
    assert render_template("{氏名} 様 {不明}", {"氏名": "山田"}) == "山田 様 {不明}"
    assert unknown_tags("", "{氏名}{不明}", ["氏名"]) == ["不明"]


def test_conditional_suffix_tag_hides_suffix_when_value_is_empty():
    template = "{氏名| 様}\n{氏名2| 様}"
    assert render_template(template, {"氏名": "山田", "氏名2": "佐藤"}) == "山田 様\n佐藤 様"
    assert render_template(template, {"氏名": "山田", "氏名2": ""}) == "山田 様\n"
    assert unknown_tags("", template, ["氏名", "氏名2"]) == []


def test_conditional_prefix_suffix_tag_hides_separator_when_value_is_empty():
    template = "{氏名A}様{、|氏名B|様}"
    assert render_template(
        template, {"氏名A": "山田", "氏名B": "佐藤"}) == "山田様、佐藤様"
    assert render_template(
        template, {"氏名A": "山田", "氏名B": ""}) == "山田様"
    assert unknown_tags("", template, ["氏名A", "氏名B"]) == []
    assert unknown_tags("", "{、|不明|様}", ["氏名A"]) == ["不明"]


def test_validate_rows_detects_invalid_and_duplicate():
    rows = [{"mail": "a@example.jp"}, {"mail": "a@example.jp"}, {"mail": "bad"}]
    errors = validate_rows(rows, "mail")
    assert 0 not in errors
    assert errors[1] == ["宛先が2行目と重複しています"]
    assert 2 in errors


def test_validate_rows_uses_actual_row_numbers_when_filtered():
    # 絞り込みで1・9・10行目だけを渡した場合、重複メッセージは
    # 渡した並び順ではなく実際のExcel行番号を指す。
    rows = [{"mail": "z@example.jp"}, {"mail": "a@example.jp"},
            {"mail": "a@example.jp"}]
    errors = validate_rows(rows, "mail", "", [2, 9, 10])
    assert errors == {2: ["宛先が9行目と重複しています"]}


def test_split_addresses():
    assert split_addresses("a@example.jp; b@example.jp,c@example.jp") == [
        "a@example.jp", "b@example.jp", "c@example.jp"
    ]


def test_typo_domain_suspects_flags_lookalike_domains():
    addresses = [
        "a@dokomo.ne.jp", "B@GMAI.COM", "c@docomo.ne.jp", "d@yahoo.ne.jp",
        "a@dokomo.ne.jp",
    ]
    suspects = typo_domain_suspects(addresses)
    # 正規のドメイン（docomo.ne.jp、Y!mobileのyahoo.ne.jp）は対象外。
    # 同じアドレスの重複は1件にまとめる。
    assert suspects == [("a@dokomo.ne.jp", "docomo.ne.jp"), ("B@GMAI.COM", "gmail.com")]


def test_carrier_domain_counts_groups_by_domain():
    addresses = [
        "a@docomo.ne.jp", "b@docomo.ne.jp", "c@ezweb.ne.jp",
        "d@example.co.jp", "e@gmail.com",
    ]
    assert carrier_domain_counts(addresses) == {"docomo.ne.jp": 2, "ezweb.ne.jp": 1}


def test_match_individual_attachments_by_column_value(tmp_path):
    exact = tmp_path / "山田商事.pdf"
    extra = tmp_path / "山田商事_請求書.xlsx"
    other = tmp_path / "未登録会社.pdf"
    for path in (exact, extra, other):
        path.write_bytes(b"x")
    mapping, unmatched = match_individual_attachments(
        [{"事業所名": "山田商事"}, {"事業所名": "鈴木商店"}],
        "事業所名", [str(exact), str(extra), str(other)])
    assert mapping == {0: sorted([str(exact), str(extra)])}
    assert unmatched == [str(other)]


def test_match_individual_attachments_by_two_column_values(tmp_path):
    exact = tmp_path / "12_山田商事.pdf"
    extra = tmp_path / "12_山田商事_請求書.xlsx"
    wrong = tmp_path / "13_山田商事.pdf"
    for path in (exact, extra, wrong):
        path.write_bytes(b"x")
    mapping, unmatched = match_individual_attachments(
        [{"NO.": "12", "事業所名": "山田商事"},
         {"NO.": "13", "事業所名": ""}],
        ["NO.", "事業所名"], [str(exact), str(extra), str(wrong)])
    assert mapping == {0: sorted([str(exact), str(extra)])}
    assert unmatched == [str(wrong)]
