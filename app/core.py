from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

EMAIL_RE = re.compile(r"^[^@\s,;]+@[^@\s,;]+\.[^@\s,;]+$")
TAG_RE = re.compile(r"\{([^{}]+)\}")
CONDITIONAL_TAG_RE = re.compile(r"\{([^{}|]+)\|([^{}]*)\}")
CONDITIONAL_PREFIX_SUFFIX_TAG_RE = re.compile(r"\{([^{}|]*)\|([^{}|]+)\|([^{}]*)\}")

# よくある打ち間違いドメインと、想定される正しいドメイン。
# これらは第三者が実際に取得してMXを運用している場合があり、
# 誤送信しても配信不能通知が返らず、添付ファイルごと他人に届いてしまう。
# 形式チェックでは検出できないため、名前で照合して送信前に警告する。
TYPO_DOMAINS = {
    "dokomo.ne.jp": "docomo.ne.jp",
    "docomo.co.jp": "docomo.ne.jp",
    "docomo.jp": "docomo.ne.jp",
    "docomo.ne.jo": "docomo.ne.jp",
    "docmo.ne.jp": "docomo.ne.jp",
    "docoomo.ne.jp": "docomo.ne.jp",
    "dcomo.ne.jp": "docomo.ne.jp",
    "ezwed.ne.jp": "ezweb.ne.jp",
    "ezweb.ne.jo": "ezweb.ne.jp",
    "ezweb.co.jp": "ezweb.ne.jp",
    "ezwbe.ne.jp": "ezweb.ne.jp",
    "i.softbank.ne.jp": "i.softbank.jp",
    "i.softbank.co.jp": "i.softbank.jp",
    "softbank.jp": "softbank.ne.jp",
    "softbnak.ne.jp": "softbank.ne.jp",
    "gmai.com": "gmail.com",
    "gmial.com": "gmail.com",
    "gmali.com": "gmail.com",
    "gmaill.com": "gmail.com",
    "gnail.com": "gmail.com",
    "gmail.con": "gmail.com",
    "gmail.co.jp": "gmail.com",
    "gmail.ne.jp": "gmail.com",
    "yaho.co.jp": "yahoo.co.jp",
    "yahho.co.jp": "yahoo.co.jp",
    "yahoo.com.jp": "yahoo.co.jp",
    "yahoo.co.jo": "yahoo.co.jp",
    "yahoo.jp": "yahoo.co.jp",
    "iclod.com": "icloud.com",
    "icoud.com": "icloud.com",
    "iclould.com": "icloud.com",
    "icloud.co.jp": "icloud.com",
    "hotmai.com": "hotmail.com",
    "hotmial.com": "hotmail.com",
    "homail.com": "hotmail.com",
    "outlook.co.jp": "outlook.jp",
    "outolook.jp": "outlook.jp",
    "nifty.co.jp": "nifty.com",
}

# 携帯キャリアのメールドメイン。「なりすまし規制」「パソコンメール拒否」は
# 送信側にエラーを返さずに破棄するため、不達に気づけない。
CARRIER_DOMAINS = {
    "docomo.ne.jp", "ezweb.ne.jp", "au.com", "i.softbank.jp", "softbank.ne.jp",
    "vodafone.ne.jp", "disney.ne.jp", "ymobile.ne.jp", "y-mobile.ne.jp",
    "willcom.com", "emnet.ne.jp", "pdx.ne.jp", "rakuten.jp",
}


@dataclass
class ImportResult:
    headers: list[str]
    rows: list[dict[str, str]]
    warnings: list[str]


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _unique_headers(values: Iterable[object]) -> tuple[list[str], list[str]]:
    headers, warnings, seen = [], [], {}
    for index, value in enumerate(values, 1):
        base = _text(value) or f"列{index}"
        seen[base] = seen.get(base, 0) + 1
        name = base if seen[base] == 1 else f"{base}_{seen[base]}"
        if name != base:
            warnings.append(f"重複ヘッダー「{base}」を「{name}」に変更しました。")
        headers.append(name)
    return headers, warnings


def load_recipient_file(path: str) -> ImportResult:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".csv":
        raw = None
        for encoding in ("utf-8-sig", "cp932", "utf-8"):
            try:
                with source.open(encoding=encoding, newline="") as handle:
                    raw = list(csv.reader(handle))
                break
            except UnicodeDecodeError:
                continue
        if raw is None:
            raise ValueError("CSVの文字コードを判定できませんでした。")
    elif suffix == ".xlsx":
        book = load_workbook(source, read_only=True, data_only=True)
        sheet = book.active
        raw = [list(row) for row in sheet.iter_rows(values_only=True)]
        book.close()
    elif suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("xlsの読込には xlrd が必要です。") from exc
        sheet = xlrd.open_workbook(source).sheet_by_index(0)
        raw = [sheet.row_values(i) for i in range(sheet.nrows)]
    else:
        raise ValueError("対応形式は xlsx、xls、csv です。")
    if not raw:
        raise ValueError("ファイルにデータがありません。")
    headers, warnings = _unique_headers(raw[0])
    rows = []
    for values in raw[1:]:
        values = list(values) + [""] * (len(headers) - len(values))
        row = {header: _text(values[i]) for i, header in enumerate(headers)}
        if any(row.values()):
            rows.append(row)
    if not rows:
        warnings.append("データ行がありません。")
    return ImportResult(headers, rows, warnings)


def export_recipient_file(path: str, headers: list[str], rows: list[dict[str, str]]) -> None:
    book = Workbook()
    sheet = book.active
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    book.save(path)


def split_addresses(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[,;\n]", value or "") if item.strip()]


def is_valid_email(value: str) -> bool:
    return bool(EMAIL_RE.fullmatch(value.strip()))


def domain_of(address: str) -> str:
    return address.strip().rpartition("@")[2].casefold()


def typo_domain_suspects(addresses: Iterable[str]) -> list[tuple[str, str]]:
    """打ち間違いが疑われるアドレスと、想定される正しいドメインを返す。"""
    suspects, seen = [], set()
    for address in addresses:
        value = address.strip()
        correction = TYPO_DOMAINS.get(domain_of(value))
        if correction and value.casefold() not in seen:
            seen.add(value.casefold())
            suspects.append((value, correction))
    return suspects


def carrier_domain_counts(addresses: Iterable[str]) -> dict[str, int]:
    """携帯キャリア宛の件数をドメインごとに数え、多い順に返す。"""
    counts: dict[str, int] = {}
    for address in addresses:
        domain = domain_of(address)
        if domain in CARRIER_DOMAINS:
            counts[domain] = counts.get(domain, 0) + 1
    return dict(sorted(counts.items(), key=lambda item: (-item[1], item[0])))


def validate_rows(rows: list[dict[str, str]], to_column: str,
                  cc_column: str = "",
                  row_numbers: list[int] | None = None) -> dict[int, list[str]]:
    """rowsを検査し、rowsの位置をキーにしたエラー一覧を返す。

    絞り込み後の一部だけを渡す場合、row_numbersに各行の実際の
    Excel / CSV行番号を渡す。省略すると先頭行を2行目として数える。
    重複メッセージの行番号がrows内の並び順基準になると、
    絞り込みの有無で文言が変わり、確認済みの許可が外れてしまう。
    """
    errors: dict[int, list[str]] = {}
    seen: dict[str, int] = {}
    for index, row in enumerate(rows):
        line = row_numbers[index] if row_numbers else index + 2
        issues = []
        recipients = split_addresses(row.get(to_column, ""))
        if not recipients:
            issues.append("宛先が空です")
        elif any(not is_valid_email(address) for address in recipients):
            issues.append("宛先の形式が不正です")
        for address in recipients:
            key = address.casefold()
            if key in seen:
                issues.append(f"宛先が{seen[key]}行目と重複しています")
            else:
                seen[key] = line
        if cc_column:
            cc = split_addresses(row.get(cc_column, ""))
            if any(not is_valid_email(address) for address in cc):
                issues.append("CCの形式が不正です")
        if issues:
            errors[index] = issues
    return errors


def render_template(template: str, row: dict[str, str]) -> str:
    # {前置文字|列名|後置文字} は値がある場合だけ「前置＋値＋後置」を出力する。
    # 例: {、|氏名2|様} → "、佐藤様" または空文字
    template = CONDITIONAL_PREFIX_SUFFIX_TAG_RE.sub(
        lambda match: (
            match.group(1) + row.get(match.group(2), "") + match.group(3)
            if row.get(match.group(2), "") else ""
        ),
        template,
    )
    # {列名|後置文字} は値がある場合だけ「値＋後置文字」を出力する。
    # 例: {氏名2| 様} → "佐藤 様" または空文字
    template = CONDITIONAL_TAG_RE.sub(
        lambda match: (
            row.get(match.group(1), "") + match.group(2)
            if row.get(match.group(1), "") else ""
        ),
        template,
    )
    return TAG_RE.sub(lambda match: row.get(match.group(1), match.group(0)), template)


def unknown_tags(subject: str, body: str, headers: list[str]) -> list[str]:
    source = subject + "\n" + body
    tags = set()
    for tag in TAG_RE.findall(source):
        parts = tag.split("|")
        tags.add(parts[1] if len(parts) >= 3 else parts[0])
    return sorted(tags.difference(headers))


def match_individual_attachments(
        rows: list[dict[str, str]], match_columns: str | list[str] | tuple[str, ...],
        file_paths: list[str]) -> tuple[dict[int, list[str]], list[str]]:
    """1～複数列の値を「_」でつないでファイル名と照合する。

    「NO.」が「12」、「事業所名」が「山田商事」なら、
    12_山田商事.pdf、12_山田商事_請求書.pdf などが一致する。
    """
    columns = [match_columns] if isinstance(match_columns, str) else list(match_columns)
    result: dict[int, list[str]] = {}
    used: set[str] = set()
    for index, row in enumerate(rows):
        values = [row.get(column, "").strip() for column in columns]
        if not columns or any(not value for value in values):
            continue
        key = "_".join(values)
        matches = []
        for file_path in file_paths:
            stem = Path(file_path).stem.strip()
            if stem == key or any(
                    stem.startswith(key + separator)
                    for separator in ("_", "-", "－", " ", "　")):
                matches.append(file_path)
                used.add(file_path)
        if matches:
            result[index] = sorted(matches)
    unmatched = sorted(path for path in file_paths if path not in used)
    return result, unmatched


# 検索時の表記ゆれをそろえる。NFKC で全角・半角を統一し、ふりがなとして
# 入力されやすいひらがなはカタカナに寄せる。
def normalize_search_text(value: object) -> str:
    normalized = unicodedata.normalize("NFKC", str(value)).casefold()
    return "".join(
        chr(ord(char) + 0x60) if "ぁ" <= char <= "ゖ" else char
        for char in normalized
    )
